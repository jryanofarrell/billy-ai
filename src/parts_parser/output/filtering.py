from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from parts_parser.models import PartRecord


HEADER_NAMES = {
    "PART",
    "PARTNO",
    "PARTNUMBER",
    "PARTNUM",
    "ITEM",
    "ITEMNO",
    "ITEMNUMBER",
}


class OutputError(Exception):
    """An error that prevents output processing from completing."""


def normalize_key(raw: str) -> str:
    allowed = "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
    return "".join(character for character in raw.upper() if character in allowed)


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass
class FilterEntry:
    raw: str
    normalized: str
    row: int


@dataclass
class FilterSheet:
    path: Path
    column_label: str
    entries: list[FilterEntry]


def load_filter_sheet(path: Path) -> FilterSheet:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise OutputError(
            "The Excel file could not be opened. Check that it is a valid Excel file and try again."
        ) from error

    try:
        worksheet = workbook.worksheets[0]
        selected_column = 1
        data_start_row = 1
        column_label = "column A (no header found)"

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1), ())
        for cell in header_row:
            header = _cell_to_text(cell.value)
            if normalize_key(header) in HEADER_NAMES:
                selected_column = cell.column
                data_start_row = 2
                column_letter = get_column_letter(selected_column)
                column_label = f'column {column_letter} ("{header}")'
                break

        entries = []
        for row in range(data_start_row, worksheet.max_row + 1):
            raw = _cell_to_text(worksheet.cell(row=row, column=selected_column).value)
            if raw:
                entries.append(FilterEntry(raw=raw, normalized=normalize_key(raw), row=row))
    finally:
        workbook.close()

    if not entries:
        raise OutputError(
            "No part numbers were found in the Excel file. Expected them in a column "
            "headed 'Part No' (or in the first column)."
        )

    return FilterSheet(path=path, column_label=column_label, entries=entries)


@dataclass
class MatchResult:
    filter_raw: str
    match_type: str
    matched_part_nos: list[str]
    note: str


@dataclass
class MatchReport:
    column_label: str
    results: list[MatchResult]


def match_parts(
    filter_sheet: FilterSheet, parts: list[PartRecord]
) -> tuple[list[PartRecord], MatchReport]:
    exact_index: dict[str, PartRecord] = {}
    norm_index: dict[str, list[str]] = {}
    for part in parts:
        exact_index.setdefault(part.part_no, part)
        candidates = norm_index.setdefault(normalize_key(part.part_no), [])
        if part.part_no not in candidates:
            candidates.append(part.part_no)

    results = []
    matched_part_nos: set[str] = set()
    for entry in filter_sheet.entries:
        note = ""
        if entry.raw in exact_index:
            match_type = "exact"
            matched = [entry.raw]
        elif entry.normalized in norm_index:
            matched = norm_index[entry.normalized]
            if len(matched) == 1:
                match_type = "normalized"
            else:
                match_type = "collision"
                note = (
                    f"{len(matched)} different source part numbers normalize to the same value"
                )
        else:
            match_type = "unmatched"
            matched = []

        matched_part_nos.update(matched)
        results.append(
            MatchResult(
                filter_raw=entry.raw,
                match_type=match_type,
                matched_part_nos=list(matched),
                note=note,
            )
        )

    matched_parts = []
    added_part_nos: set[str] = set()
    for part in parts:
        if part.part_no in matched_part_nos and part.part_no not in added_part_nos:
            matched_parts.append(part)
            added_part_nos.add(part.part_no)

    return matched_parts, MatchReport(column_label=filter_sheet.column_label, results=results)
