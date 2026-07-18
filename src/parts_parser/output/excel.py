from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from parts_parser.models import PartRecord
from parts_parser.output.filtering import MatchReport


PDF_COLUMNS = [
    "Part No",
    "Category",
    "Subcategory",
    "Series",
    "Size",
    "Sequence",
    "Page",
]
WEB_FIXED_COLUMNS = ["Part No", "Category", "Subcategory", "Series", "Size"]


def _write_headers(worksheet: Worksheet, headers: list[str], *, row: int = 1) -> None:
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=column, value=header)
        cell.font = Font(bold=True)


def _write_pdf_parts_sheet(worksheet: Worksheet, parts: list[PartRecord]) -> None:
    _write_headers(worksheet, PDF_COLUMNS)
    worksheet.freeze_panes = "A2"

    for row, part in enumerate(parts, start=2):
        values: list[str | int | None] = [
            part.part_no,
            part.category,
            part.subcategory,
            part.series,
            part.description,
            part.sequence,
            part.page_number,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=column, value=value)


def _write_web_parts_sheet(worksheet: Worksheet, parts: list[PartRecord]) -> None:
    attribute_labels = sorted({label for part in parts for label in part.attributes})
    _write_headers(worksheet, WEB_FIXED_COLUMNS + attribute_labels)
    worksheet.freeze_panes = "A2"

    for row, part in enumerate(parts, start=2):
        values: list[str | int | None] = [
            part.part_no,
            part.category,
            part.subcategory,
            part.series,
            part.description,
        ]
        values.extend(part.attributes.get(label, "") for label in attribute_labels)
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=column, value=value)


def _write_match_report(worksheet: Worksheet, match_report: MatchReport) -> None:
    title = worksheet.cell(
        row=1,
        column=1,
        value=f"Part numbers read from {match_report.column_label}",
    )
    title.font = Font(italic=True)
    _write_headers(
        worksheet,
        ["Filter Value", "Match Type", "Matched Part No", "Note"],
        row=2,
    )

    for row, result in enumerate(match_report.results, start=3):
        values = [
            result.filter_raw,
            result.match_type,
            ", ".join(result.matched_part_nos),
            result.note,
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=column, value=value)


def write_workbook(
    parts: list[PartRecord],
    out_path: Path,
    *,
    mode: Literal["web", "pdf"],
    match_report: MatchReport | None = None,
) -> None:
    workbook = Workbook()
    parts_sheet = workbook.active
    parts_sheet.title = "Parts"
    if mode == "pdf":
        _write_pdf_parts_sheet(parts_sheet, parts)
    else:
        _write_web_parts_sheet(parts_sheet, parts)

    if match_report is not None:
        report_sheet = workbook.create_sheet("Match Report")
        _write_match_report(report_sheet, match_report)

    workbook.save(out_path)
