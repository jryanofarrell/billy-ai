from openpyxl import load_workbook

from parts_parser.models import PartRecord
from parts_parser.output.excel import PDF_COLUMNS, WEB_FIXED_COLUMNS, write_workbook
from parts_parser.output.filtering import MatchReport, MatchResult


def test_write_workbook_pdf_has_exact_headers_and_preserves_part_number(tmp_path):
    output_path = tmp_path / "parts.xlsx"
    parts = [
        PartRecord(
            part_no="007345",
            category="Synthetic category",
            subcategory="Synthetic subcategory",
            series="Synthetic series",
            description="Synthetic description",
            sequence=1,
        )
    ]

    write_workbook(parts, output_path, mode="pdf")

    workbook = load_workbook(output_path)
    try:
        worksheet = workbook["Parts"]
        assert [cell.value for cell in worksheet[1]] == PDF_COLUMNS
        assert worksheet["A2"].value == "007345"
        assert isinstance(worksheet["A2"].value, str)
        assert workbook.sheetnames == ["Parts"]
    finally:
        workbook.close()


def test_write_workbook_web_has_sorted_attribute_union_and_blank_missing_value(
    tmp_path,
):
    output_path = tmp_path / "parts.xlsx"
    parts = [
        PartRecord(part_no="A-1", attributes={"Size": "1 in", "Color": "Blue"}),
        PartRecord(part_no="B-2", attributes={"Material": "Brass", "Size": "2 in"}),
    ]

    write_workbook(parts, output_path, mode="web")

    workbook = load_workbook(output_path)
    try:
        worksheet = workbook["Parts"]
        expected_headers = WEB_FIXED_COLUMNS + ["Color", "Material", "Size"]
        assert [cell.value for cell in worksheet[1]] == expected_headers
        material_column = expected_headers.index("Material") + 1
        assert worksheet.cell(row=2, column=material_column).value is None
        assert worksheet.cell(row=3, column=material_column).value == "Brass"
    finally:
        workbook.close()


def test_write_workbook_adds_match_report_with_column_note_and_every_result(
    tmp_path,
):
    output_path = tmp_path / "parts.xlsx"
    report = MatchReport(
        column_label='column B ("Part Number")',
        results=[
            MatchResult(
                filter_raw="A 1",
                match_type="normalized",
                matched_part_nos=["A-1"],
                note="",
            ),
            MatchResult(
                filter_raw="MISSING",
                match_type="unmatched",
                matched_part_nos=[],
                note="Not found",
            ),
        ],
    )

    write_workbook(
        [PartRecord(part_no="A-1")],
        output_path,
        mode="pdf",
        match_report=report,
    )

    workbook = load_workbook(output_path)
    try:
        assert workbook.sheetnames == ["Parts", "Match Report"]
        worksheet = workbook["Match Report"]
        assert worksheet["A1"].value == 'Part numbers read from column B ("Part Number")'
        assert [cell.value for cell in worksheet[2]] == [
            "Filter Value",
            "Match Type",
            "Matched Part No",
            "Note",
        ]
        assert worksheet.max_row == 2 + len(report.results)
        assert [worksheet.cell(row=row, column=1).value for row in (3, 4)] == [
            "A 1",
            "MISSING",
        ]
    finally:
        workbook.close()
